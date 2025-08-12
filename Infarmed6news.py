from datetime import datetime
import os
import re
import logging
from typing import Dict, List, Optional, Tuple, Any, Generator
from urllib.parse import urljoin
import fasttext
import scrapy
from deep_translator import GoogleTranslator
from langcodes import Language as Lang
from langdetect import DetectorFactory, detect
from sumy.nlp.tokenizers import Tokenizer
from sumy.parsers.plaintext import PlaintextParser
from sumy.summarizers.lsa import LsaSummarizer
from scrapy.http import Response
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from scrapy.crawler import CrawlerProcess
import re

# Initialize language detection
DetectorFactory.seed = 0

class ExcelExporter:
    """Handles Excel export using OpenPyXL."""
    
    def __init__(self, filename='Infarmednews_items.xlsx'):
        self.filename = filename
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "News Items"
        self.row_counter = 1
        self._setup_headers()
        
    def _setup_headers(self):
        """Set up the header row with formatting."""
        headers = [
            'Title',
            'Summary',
            'Article URL',
            'Date',
            'Document_Type',
            'Product_Type',
            'Countries',
            'Regions',
            'Drug_names',
            'Language',
            'Source URL'
        ]

        
        # Apply header styling
        bold_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        for col_num, header in enumerate(headers, 1):
            cell = self.sheet.cell(row=1, column=col_num, value=header)
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            
            # Set initial column width
            self.sheet.column_dimensions[get_column_letter(col_num)].width = 20
        
        self.row_counter += 1
        
    def add_item(self, item):
        """Add a news item to the Excel sheet."""
        values = [
            item.get('title', ''),
            item.get('summary', ''),
            item.get('article_url', ''),
            item.get('date', ''),
            item.get('document_type', ''),
            item.get('product_type', ''),
            item.get('countries', ''),
            item.get('regions', ''),
            item.get('drug_names', ''),
            item.get('language', ''),
            item.get('source_url', '')
        ]

        
        thin_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))
        
        for col_num, value in enumerate(values, 1):
            cell = self.sheet.cell(row=self.row_counter, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
            
            # Auto-adjust column width for long text
            col_letter = get_column_letter(col_num)
            if len(str(value)) > 30:
                self.sheet.column_dimensions[col_letter].width = 30
        
        self.row_counter += 1
        
    def save(self):
        """Save the workbook to file."""
        try:
            self.workbook.save(self.filename)
            logging.info(f"Successfully saved Excel file: {self.filename}")
        except Exception as e:
            logging.error(f"Failed to save Excel file: {str(e)}")
            raise

DetectorFactory.seed = 0

class DocumentClassifier:
    def __init__(self):
        self._initialize_drug_lookup()
        self.document_types = self._get_document_types()
        self.product_types = self._get_product_types()


    def _initialize_drug_lookup(self):
                # ✅ LOAD drug terms from .tsv file
        tsv_path = 'https://raw.githubusercontent.com/MariaKlap/Drug-Name-Database/refs/heads/main/drug.target.interaction.tsv'  
        

        try:
            df = pd.read_csv(tsv_path, sep='\t')
        except UnicodeDecodeError:
            df = pd.read_csv(tsv_path, sep='\t', encoding='ISO-8859-1')
        print(f"📊 DRUG_NAME column row count (including duplicates and empty): {len(df['DRUG_NAME'])}")

        # Limit to specific columns only
        allowed_columns = {'DRUG_NAME', 'SWISSPROT', 'ACTION_TYPE', 'TARGET_CLASS', 'TARGET_NAME'}
        allowed_columns = [col for col in df.columns if col in allowed_columns]

        terms = set()
        for col in allowed_columns:
            col_terms = df[col].dropna().astype(str)
            col_terms = {t.strip().lower() for t in col_terms if len(t.strip()) > 3}
            terms.update(col_terms)

        self.drug_terms_set = terms
        print(f"✅ Loaded {len(self.drug_terms_set)} drug terms from TSV columns: {', '.join(allowed_columns)}")

    def classify_document(self, text: str) -> Dict[str, str]:
        """Classify document type based on English text content."""
        if not text:
            return {
                'document_type': 'Other Type', 
                'matched_keywords': 'empty text'
            }
            
        text_lower = text.lower()
        for doc_type, keywords in self.document_types.items():
            matched = [
                kw for kw in keywords 
                if re.search(rf'\b{re.escape(kw)}\b', text_lower)
            ]
            if matched:
                return {
                    'document_type': doc_type,
                    'matched_keywords': ", ".join(matched[:3])
                }
        return {
            'document_type': 'Other Type', 
            'matched_keywords': 'unclassified'
        }
    
    def classify_product(self, text: str) -> Dict[str, Optional[str]]:
        """Classify product type from English text."""
        if not text:
            return {
                'product_type': 'Other',
                'product_keywords': 'empty text',
                'drug_names': 'None'  # Changed from None to 'None'
            }
            
        text_lower = text.lower()
        drug_names = self.extract_drug_names(text)
        
        # First check specific product types (excluding Drug Product)
        for product_type, keywords in self.product_types.items():
            if product_type == 'Drug Product':
                continue
                
            matched = [
                kw for kw in keywords 
                if re.search(rf'\b{re.escape(kw)}\b', text_lower)
            ]
            if matched:
                return {
                    'product_type': product_type,
                    'product_keywords': ", ".join(matched[:3]),
                    'drug_names': ", ".join(drug_names) if drug_names != 'None' else 'None'  # Changed handling here
                }

        # Default to Drug Product if drug names found
        if drug_names != 'None':  # Changed condition here
            return {
                'product_type': 'Drug Product',
                'product_keywords': f"Identified drug names: {', '.join(drug_names[:3])}",
                'drug_names': ", ".join(drug_names)
            }

        # Default case
        return {
            'product_type': 'Other',
            'product_keywords': 'unclassified',
            'drug_names': 'None'  # Changed from None to 'None'
        }


    def extract_drug_names(self, text: str) -> List[str]:
        if not text.strip():
            return []

        matched = []
        for term in self.drug_terms_set:
            # Escape regex special chars and match as whole word/phrase
            pattern = r'\b' + re.escape(term) + r'\b'
            if re.search(pattern, text, flags=re.IGNORECASE):
                matched.append(term)
        return matched

        return sorted(found_terms) if found_terms else 'None'  # Changed from None to 'None'
        
    def _is_likely_drug_name(self, text: str) -> bool:
        """Very strict drug name detection with multiple validation checks."""
        if len(text) < 4:  # Increased minimum length
            return False
        
        text_lower = text.lower()
        
        if any(c.isdigit() for c in text) and not re.search(r'\d+[A-Za-z]', text):
            return False
        
        drug_suffixes = {
        'mab', 'nib', 'vir', 'pib', 'caine', 'pril', 'oxetine', 'olol',
        'azepam', 'oxacin', 'tide', 'barb', 'mycin', 'cycline', 'sartan',
        'prazole', 'triptan', 'glitazar', 'vastatin', 'profen', 'lukast'
        }
        
        has_drug_suffix = any(text_lower.endswith(suffix) for suffix in drug_suffixes)
        
        is_brand_name = (
            re.match(r'^[A-Z][a-z]+$', text) and 
            len(text) >= 5 and
            not text_lower in {'patient', 'study', 'clinical'}
            )
        
        is_generic_name = (
            re.match(r'^[A-Z][a-z]+-[A-Z][a-z]+$', text) or
            re.match(r'^[A-Z][a-z]+[A-Z][a-z]+$', text)
            )
        
        criteria_met = 0
        if has_drug_suffix:
            criteria_met += 1
        if is_brand_name:
            criteria_met += 1
        if is_generic_name:
            criteria_met += 1
            
        return criteria_met >= 2


    def _get_document_types(self) -> Dict[str, List[str]]:
        """Return document type classification dictionary with improved keywords."""
        return {
            'Announcement': ['announcement', 'notification', 'bulletin'],
            'Expert Report': ['expert report', 'technical report', 'scientific opinion'],
            'Amendment': ['amendment', 'regulation change', 'regulatory update'],
            'Law': ['law', 'legislation', 'statute'],
            'Directive': ['directive', 'guideline', 'policy'],
            'Order': ['order', 'decision', 'ruling', 'decree'],
            'Information Note': ['information note', 'information bulletin', 'notice'],
            'Q&A': ['questions and answers', 'q&a', 'faq', 'frequently asked'],
            'Instructions': ['instructions', 'manual', 'guidance', 'procedure'],
            'Resolution': ['resolution', 'conclusion', 'determination'],
            'Consultation': ['consultation', 'public hearing', 'stakeholder input'],
            'Product Info': ['product information', 'package leaflet', 'product update'],
            'Regulatory Decision': ['regulatory decision', 'approval summary', 'assessment'],
            'Evaluation': ['evaluation report', 'assessment report', 'review report'],
            'Recommendation': ['recommendation', 'advice', 'suggestion'],
            'Checklist': ['checklist', 'verification list', 'review points'],
            'Approval Tracker': ['approval tracker', 'authorization status', 'timeline'],
            'CHMP Opinion': ['chmp opinion', 'committee opinion', 'scientific opinion'],
            'Committee': ['committee', 'working group', 'task force'],
            'CV': ['curriculum vitae', 'cv', 'resume'],
            'EPAR': ['epar', 'european public assessment report'],
            'Letter': ['letter', 'correspondence', 'official communication'],
            'Meeting': ['meeting', 'conference', 'session'],
            'Withdrawal': ['withdrawn application', 'cancelled submission'],
            'Communication': ['communication', 'announcement', 'message'],
            'Decree': ['decree', 'royal decree', 'official order'],
            'Form': ['form', 'application form', 'submission form'],
            'Regulatory History': ['regulatory history', 'dossier history', 'timeline'],
            'Press Release': ['press release', 'news release', 'media statement'],
            'Ordinance': ['ordinance', 'local regulation', 'municipal law'],
            'Advisory Committee': ['advisory committee', 'committee profile'],
            'Voting': ['voting', 'committee vote', 'decision outcome'],
            'Petition': ['citizen petition', 'public petition', 'request'],
            'Federal Register': ['federal register', 'official gazette', 'journal'],
            'Inspection': ['inspection report', 'audit report', 'site visit'],
            'SOP': ['sop', 'standard procedure', 'operating protocol'],
            'BLA Approval': ['bla', 'biologics license application'],
            'BLA Supplement': ['supplemental bla', 'bla amendment'],
            'NDA Supplement': ['supplemental nda', 'nda amendment'],
            '510(k)': ['510(k)', 'premarket notification'],
            'NDA Approval': ['new drug application'],
            'Other Type': []
        }

    def _get_product_types(self) -> Dict[str, List[str]]:
        """Return product type classification dictionary with improved keywords."""
        return {
            'Biological': [
                'biological', 'biologic', 'monoclonal antibody', 'mab', 
                'recombinant', 'cell therapy', 'gene therapy', 'blood product', 
                'plasma derived', 'therapeutic protein', 'insulin', 'erythropoietin',
                'immunoglobulin', 'stem cell'
            ],
            'Drug Product': [
                'drug product', 'finished product', 'formulation', 'dosage form',
                'tablet', 'capsule', 'injection', 'solution', 'suspension', 
                'biopharmaceutical', 'biosimilar', 'cream', 'ointment', 'gel', 
                'suppository', 'inhalation'
            ],
            'Drug Substance': [
                'drug substance', 'active substance', 'api', 'active ingredient',
                'bulk drug', 'chemical entity', 'reference standard'
            ],
            'Vaccine': [
                'vaccine', 'vaccination', 'immunization', 'antigen', 'adjuvant',
                'mmr', 'dtap', 'hpv', 'influenza', 'covid-19', 'sars-cov-2'
            ],
            'Small Molecule': [
                'small molecule', 'chemical drug', 'synthetic', 'organic compound',
                'nme', 'new molecular entity', 'low molecular weight'
            ],
            'Medical Device': [
                'medical device', 'implant', 'stent', 'catheter', 'prosthesis',
                'pacemaker', 'defibrillator', 'surgical instrument'
            ],
            'IVD': [
                'ivd', 'in vitro diagnostic', 'diagnostic test', 'assay',
                'reagent', 'test kit', 'analyzer', 'rapid test'
            ],
            'Other': []
        }

    def translate_text(self, text: str, source_lang: str) -> str:
        """Translate text to English using Google Translate API with improved error handling."""
        if not text or not source_lang or source_lang.lower() == 'english':
            return text
            
        try:
            # Handle language code mapping
            lang_code_map = {
                'portuguese': 'pt',
                'chinese': 'zh',
                'arabic': 'ar',
                'russian': 'ru'
            }
            
            lang_code = lang_code_map.get(source_lang.lower())
            if not lang_code:
                lang_obj = Lang(source_lang)
                lang_code = lang_obj.part1 if lang_obj else None
            
            if not lang_code:
                logging.warning(f"No ISO639-1 code for language: {source_lang}")
                return text
            
            # Split long text to avoid API limits
            if len(text) > 5000:
                parts = [text[i:i+4000] for i in range(0, len(text), 4000)]
                translated_parts = []
                for part in parts:
                    translated = GoogleTranslator(source=lang_code, target='en').translate(part)
                    if translated:
                        translated_parts.append(translated)
                return " ".join(translated_parts)
            
            return GoogleTranslator(source=lang_code, target='en').translate(text)
        except Exception as e:
            logging.warning(f"Translation failed ({source_lang}): {str(e)}")
            return text

class DocumentProcessor:
    def __init__(self):
        # Initialize translation with deep_translator
        self.translator = GoogleTranslator
        # ... rest of your initialization code
        
    def translate_to_english(self, text: str) -> str:
        """Translate text to English if it's not already in English."""
        if not text.strip():
            return text
            
        try:
            # First try fasttext for language detection if available
            if hasattr(self, 'fasttext_model') and self.fasttext_model:
                prediction = self.fasttext_model.predict(text.replace("\n", " "))
                lang_code = prediction[0][0].replace('__label__', '')
                if lang_code == 'en':
                    return text
            else:
                # Fallback to langdetect
                if detect(text) == 'en':
                    return text
            
            # Translate using deep_translator
            translation = GoogleTranslator(source='auto', target='en').translate(text)
            return translation if translation else text
        except Exception as e:
            logging.warning(f"Translation failed: {e}")
            return text  # Return original if translation fails

class CountryDetector:
    """Handles country and region detection in English text."""
    
    def __init__(self):
        self.country_patterns = self._load_country_patterns()
        self.region_mapping = self._load_region_mapping()
        self.language_to_country = self._load_language_mapping()
        self.country_tlds = self._load_tld_mapping()
        
        # Precompile regex patterns for faster matching
        self.country_regex = {
            country: [re.compile(rf'\b{re.escape(pattern)}\b', re.IGNORECASE) 
                     for pattern in patterns]
            for country, patterns in self.country_patterns.items()
        }
    
    def translate_to_english(self, text: str) -> str:
        """Helper method to translate text to English."""
        if not text.strip():
            return text
            
        try:
            # First try fasttext for language detection if available
            if hasattr(self, 'fasttext_model') and self.fasttext_model:
                prediction = self.fasttext_model.predict(text[:500].replace("\n", " "))
                lang_code = prediction[0][0].replace('__label__', '')
                if lang_code == 'en':
                    return text
            else:
                # Fallback to langdetect
                if detect(text[:500]) == 'en':
                    return text
            
            # Translate using deep_translator
            translation = GoogleTranslator(source='auto', target='en').translate(text)
            return translation if translation else text
        except Exception as e:
            logging.warning(f"Translation failed: {e}")
            return text

    def _load_country_patterns(self) -> Dict[str, List[str]]:
        """Load optimized country detection patterns."""
        # (Same country patterns as original, but could be optimized further)
        return {
                        'Albania': ['albania', 'shqipëria', 'tirana'],
            'Andorra': ['andorra', 'andorre'],
            'Austria': ['austria', 'österreich', 'vienna'],
            'Belarus': ['belarus', 'belarús', 'minsk'],
            'Belgium': ['belgium', 'belgique', 'belgie', 'brussels'],
            'Bosnia and Herzegovina': ['bosnia', 'herzegovina', 'sarajevo'],
            'Bulgaria': ['bulgaria', 'българия', 'sofia'],
            'Croatia': ['croatia', 'hrvatska', 'zagreb'],
            'Cyprus': ['cyprus', 'κύπρος', 'lefkosia'],
            'Czech Republic': ['czech republic', 'česko', 'prague'],
            'Denmark': ['denmark', 'danmark', 'copenhagen'],
            'Estonia': ['estonia', 'eesti', 'tallinn'],
            'Finland': ['finland', 'suomi', 'helsinki'],
            'France': ['france', 'french', 'paris'],
            'Germany': ['germany', 'deutschland', 'berlin'],
            'Greece': ['greece', 'ελλάδα', 'athens'],
            'Hungary': ['hungary', 'magyarország', 'budapest'],
            'Iceland': ['iceland', 'ísland', 'reykjavik'],
            'Ireland': ['ireland', 'éire', 'dublin'],
            'Italy': ['italy', 'italia', 'rome'],
            'Latvia': ['latvia', 'latvija', 'riga'],
            'Liechtenstein': ['liechtenstein', 'vaduz'],
            'Lithuania': ['lithuania', 'lietuva', 'vilnius'],
            'Luxembourg': ['luxembourg', 'luxemburg', 'luxembourg city'],
            'Malta': ['malta', 'valletta'],
            'Moldova': ['moldova', 'chișinău'],
            'Monaco': ['monaco', 'monaco-ville'],
            'Montenegro': ['montenegro', 'crna gora', 'podgorica'],
            'Netherlands': ['netherlands', 'nederland', 'holland', 'amsterdam'],
            'North Macedonia': ['north macedonia', 'macedonia', 'skopje'],
            'Norway': ['norway', 'norge', 'oslo'],
            'Poland': ['poland', 'polska', 'warsaw'],
            'Portugal': ['portugal', 'lisbon'],
            'Romania': ['romania', 'românia', 'bucharest'],
            'Russia': ['russia', 'россия', 'moscow'],
            'San Marino': ['san marino'],
            'Serbia': ['serbia', 'srbija', 'belgrade'],
            'Slovakia': ['slovakia', 'slovensko', 'bratislava'],
            'Slovenia': ['slovenia', 'slovenija', 'ljubljana'],
            'Spain': ['spain', 'españa', 'madrid'],
            'Sweden': ['sweden', 'sverige', 'stockholm'],
            'Switzerland': ['switzerland', 'suisse', 'schweiz', 'bern'],
            'Ukraine': ['ukraine', 'україна', 'kyiv'],
            'United Kingdom': ['uk', 'united kingdom', 'britain', 'london'],
            'Vatican City': ['vatican', 'holy see'],

            # Americas (Complete list)
            'Antigua and Barbuda': ['antigua', 'barbuda', "antigua and barbuda", 'saint john'],
            'Argentina': ['argentina', 'buenos aires', 'argentine republic'],
            'Bahamas': ['bahamas', 'nassau', 'commonwealth of the bahamas'],
            'Barbados': ['barbados', 'bridgetown'],
            'Belize': ['belize', 'belmopan'],
            'Bolivia': ['bolivia', 'sucre', 'la paz', 'plurinational state'],
            'Brazil': ['brazil', 'brasil', 'brasília', 'rio de janeiro', 'federative republic'],
            'Canada': ['canada', 'ottawa', 'toronto', 'ontario', 'quebec'],
            'Chile': ['chile', 'santiago', 'republic of chile'],
            'Colombia': ['colombia', 'bogotá', 'bogota', 'republic of colombia'],
            'Costa Rica': ['costa rica', 'san josé', 'san jose'],
            'Cuba': ['cuba', 'havana', 'republic of cuba'],
            'Dominica': ['dominica', 'roseau', 'commonwealth of dominica'],
            'Dominican Republic': ['dominican republic', 'santo domingo'],
            'Ecuador': ['ecuador', 'quito', 'republic of ecuador'],
            'El Salvador': ['el salvador', 'san salvador', 'republic of el salvador'],
            'Grenada': ['grenada', "saint george"],
            'Guatemala': ['guatemala', 'guatemala city', 'republic of guatemala'],
            'Guyana': ['guyana', 'georgetown', 'cooperative republic'],
            'Haiti': ['haiti', 'port-au-prince', 'republic of haiti'],
            'Honduras': ['honduras', 'tegucigalpa', 'republic of honduras'],
            'Jamaica': ['jamaica', 'kingston'],
            'Mexico': ['mexico', 'méxico', 'mexico city', 'cdmx', 'estados unidos mexicanos'],
            'Nicaragua': ['nicaragua', 'managua', 'republic of nicaragua'],
            'Panama': ['panama', 'panama city', 'republic of panama'],
            'Paraguay': ['paraguay', 'asunción', 'asunción', 'republic of paraguay'],
            'Peru': ['peru', 'lima', 'republic of peru'],
            'Saint Kitts and Nevis': ['saint kitts', 'nevis', 'basseterre'],
            'Saint Lucia': ['saint lucia', 'castries'],
            'Saint Vincent and the Grenadines': ['saint vincent', 'grenadines', 'kingstown'],
            'Suriname': ['suriname', 'paramaribo', 'republic of suriname'],
            'Trinidad and Tobago': ['trinidad', 'tobago', 'port of spain'],
            'United States': ['usa', 'u\\.s\\.', 'united states', 'america', 'washington dc', 'new york', 'california'],
            'Uruguay': ['uruguay', 'montevideo', 'oriental republic'],
            'Venezuela': ['venezuela', 'caracas', 'bolivarian republic'],

            # Asia (Complete list)
            'Afghanistan': ['afghanistan', 'kabul', 'islamic emirate'],
            'Armenia': ['armenia', 'yerevan', 'republic of armenia'],
            'Azerbaijan': ['azerbaijan', 'baku', 'republic of azerbaijan'],
            'Bahrain': ['bahrain', 'manama', 'kingdom of bahrain'],
            'Bangladesh': ['bangladesh', 'dhaka', "people's republic"],
            'Bhutan': ['bhutan', 'thimphu', 'kingdom of bhutan'],
            'Brunei': ['brunei', 'bandar seri begawan', 'darussalam'],
            'Cambodia': ['cambodia', 'phnom penh', 'kingdom of cambodia'],
            'China': ['china', 'zhongguo', 'beijing', 'shanghai', "people's republic"],
            'Cyprus': ['cyprus', 'nicosia', 'republic of cyprus'],
            'Georgia': ['georgia', 'tbilisi'],
            'India': ['india', 'bharat', 'new delhi', 'mumbai', 'republic of india'],
            'Indonesia': ['indonesia', 'jakarta', 'republic of indonesia'],
            'Iran': ['iran', 'tehran', 'islamic republic'],
            'Iraq': ['iraq', 'baghdad', 'republic of iraq'],
            'Israel': ['israel', 'jerusalem', 'state of israel'],
            'Japan': ['japan', 'nippon', 'tokyo'],
            'Jordan': ['jordan', 'amman', 'hashemite kingdom'],
            'Kazakhstan': ['kazakhstan', 'nur-sultan', 'astana', 'republic of kazakhstan'],
            'Kuwait': ['kuwait', 'kuwait city', 'state of kuwait'],
            'Kyrgyzstan': ['kyrgyzstan', 'bishkek', 'kyrgyz republic'],
            'Laos': ['laos', 'vientiane', "lao people's republic"],
            'Lebanon': ['lebanon', 'beirut', 'lebanese republic'],
            'Malaysia': ['malaysia', 'kuala lumpur', 'putrajaya'],
            'Maldives': ['maldives', 'malé', 'republic of maldives'],
            'Mongolia': ['mongolia', 'ulaanbaatar'],
            'Myanmar': ['myanmar', 'burma', 'naypyidaw', 'republic of the union'],
            'Nepal': ['nepal', 'kathmandu', 'federal democratic republic'],
            'North Korea': ['north korea', 'dprk', 'pyongyang', "democratic people's republic"],
            'Oman': ['oman', 'muscat', 'sultanate of oman'],
            'Pakistan': ['pakistan', 'islamabad', 'islamic republic'],
            'Palestine': ['palestine', 'ramallah', 'state of palestine'],
            'Philippines': ['philippines', 'manila', 'republic of the philippines'],
            'Qatar': ['qatar', 'doha', 'state of qatar'],
            'Russia': ['russia', 'russian federation', 'moscow'],
            'Saudi Arabia': ['saudi arabia', 'riyadh', 'kingdom of saudi arabia'],
            'Singapore': ['singapore', 'republic of singapore'],
            'South Korea': ['south korea', 'korea republic', 'seoul', 'republic of korea'],
            'Sri Lanka': ['sri lanka', 'colombo', 'sri jayawardenepura kotte'],
            'Syria': ['syria', 'damascus', 'syrian arab republic'],
            'Taiwan': ['taiwan', 'taipei', 'republic of china'],
            'Tajikistan': ['tajikistan', 'dushanbe', 'republic of tajikistan'],
            'Thailand': ['thailand', 'bangkok', 'kingdom of thailand'],
            'Timor-Leste': ['timor-leste', 'east timor', 'dili', 'democratic republic'],
            'Turkey': ['turkey', 'türkiye', 'ankara', 'republic of turkey'],
            'Turkmenistan': ['turkmenistan', 'ashgabat'],
            'United Arab Emirates': ['uae', 'united arab emirates', 'dubai', 'abu dhabi'],
            'Uzbekistan': ['uzbekistan', 'tashkent', 'republic of uzbekistan'],
            'Vietnam': ['vietnam', 'hanoi', 'socialist republic'],
            'Yemen': ['yemen', "sana'a", 'republic of yemen'],

            # Africa (Complete list)
            'Algeria': ['algeria', 'algiers', "people's democratic republic"],
            'Angola': ['angola', 'luanda', 'republic of angola'],
            'Benin': ['benin', 'porto-novo', 'republic of benin'],
            'Botswana': ['botswana', 'gaborone', 'republic of botswana'],
            'Burkina Faso': ['burkina faso', 'ouagadougou'],
            'Burundi': ['burundi', 'gitega', 'republic of burundi'],
            'Cameroon': ['cameroon', 'yaoundé', 'republic of cameroon'],
            'Cape Verde': ['cape verde', 'cabo verde', 'praia', 'republic of cape verde'],
            'Central African Republic': ['central african republic', 'bangui'],
            'Chad': ['chad', "n'djamena", 'republic of chad'],
            'Comoros': ['comoros', 'moroni', 'union of the comoros'],
            'Congo (Brazzaville)': ['republic of the congo', 'congo-brazzaville', 'brazzaville'],
            'Congo (Kinshasa)': ['democratic republic of the congo', 'drc', 'kinshasa'],
            "Côte d'Ivoire": ["côte d'ivoire", 'ivory coast', 'yamoussoukro'],
            'Djibouti': ['djibouti', 'republic of djibouti'],
            'Egypt': ['egypt', 'cairo', 'arab republic of egypt'],
            'Equatorial Guinea': ['equatorial guinea', 'malabo', 'republic of equatorial guinea'],
            'Eritrea': ['eritrea', 'asmara', 'state of eritrea'],
            'Eswatini': ['eswatini', 'swaziland', 'mbabane', 'kingdom of eswatini'],
            'Ethiopia': ['ethiopia', 'addis ababa', 'federal democratic republic'],
            'Gabon': ['gabon', 'libreville', 'gabonese republic'],
            'Gambia': ['gambia', 'banjul', 'republic of the gambia'],
            'Ghana': ['ghana', 'accra', 'republic of ghana'],
            'Guinea': ['guinea', 'conakry', 'republic of guinea'],
            'Guinea-Bissau': ['guinea-bissau', 'bissau', 'republic of guinea-bissau'],
            'Kenya': ['kenya', 'nairobi', 'republic of kenya'],
            'Lesotho': ['lesotho', 'maseru', 'kingdom of lesotho'],
            'Liberia': ['liberia', 'monrovia', 'republic of liberia'],
            'Libya': ['libya', 'tripoli', 'state of libya'],
            'Madagascar': ['madagascar', 'antananarivo', 'republic of madagascar'],
            'Malawi': ['malawi', 'lilongwe', 'republic of malawi'],
            'Mali': ['mali', 'bamako', 'republic of mali'],
            'Mauritania': ['mauritania', 'nouakchott', 'islamic republic'],
            'Mauritius': ['mauritius', 'port louis', 'republic of mauritius'],
            'Morocco': ['morocco', 'rabat', 'kingdom of morocco'],
            'Mozambique': ['mozambique', 'maputo', 'republic of mozambique'],
            'Namibia': ['namibia', 'windhoek', 'republic of namibia'],
            'Niger': ['niger', 'niamey', 'republic of niger'],
            'Nigeria': ['nigeria', 'abuja', 'federal republic of nigeria'],
            'Rwanda': ['rwanda', 'kigali', 'republic of rwanda'],
            'Sao Tome and Principe': ['são tomé and príncipe', 'sao tome', 'são tomé'],
            'Senegal': ['senegal', 'dakar', 'republic of senegal'],
            'Seychelles': ['seychelles', 'victoria', 'republic of seychelles'],
            'Sierra Leone': ['sierra leone', 'freetown', 'republic of sierra leone'],
            'Somalia': ['somalia', 'mogadishu', 'federal republic of somalia'],
            'South Africa': ['south africa', 'pretoria', 'cape town', 'republic of south africa'],
            'South Sudan': ['south sudan', 'juba', 'republic of south sudan'],
            'Sudan': ['sudan', 'khartoum', 'republic of the sudan'],
            'Tanzania': ['tanzania', 'dodoma', 'united republic of tanzania'],
            'Togo': ['togo', 'lomé', 'togolese republic'],
            'Tunisia': ['tunisia', 'tunis', 'republic of tunisia'],
            'Uganda': ['uganda', 'kampala', 'republic of uganda'],
            'Zambia': ['zambia', 'lusaka', 'republic of zambia'],
            'Zimbabwe': ['zimbabwe', 'harare', 'republic of zimbabwe'],

            # International/Regional
            'European Union': ['eu', 'european union', 'e\\.u\\.', 'brussels eu'],
            'African Union': ['african union', 'au', 'addis ababa'],
            'ASEAN': ['asean', 'southeast asia', 'jakarta'],
            'Global': ['who', 'world health organization', 'united nations', 'un', 'international'],
        }

    def _load_region_mapping(self) -> Dict[str, str]:
        """Load optimized country to region mapping."""
        # (Same as original)
        return {
            'Albania': 'Southern Europe',
            'Andorra': 'Southern Europe',
            'Austria': 'Central Europe',
            'Belarus': 'Eastern Europe',
            'Belgium': 'Western Europe',
            'Bosnia and Herzegovina': 'Southern Europe',
            'Bulgaria': 'Eastern Europe',
            'Croatia': 'Southern Europe',
            'Cyprus': 'Southern Europe',
            'Czech Republic': 'Central Europe',
            'Denmark': 'Northern Europe',
            'Estonia': 'Northern Europe',
            'Finland': 'Northern Europe',
            'France': 'Western Europe',
            'Germany': 'Central Europe',
            'Greece': 'Southern Europe',
            'Hungary': 'Central Europe',
            'Iceland': 'Northern Europe',
            'Ireland': 'Northern Europe',
            'Italy': 'Southern Europe',
            'Latvia': 'Northern Europe',
            'Liechtenstein': 'Central Europe',
            'Lithuania': 'Northern Europe',
            'Luxembourg': 'Western Europe',
            'Malta': 'Southern Europe',
            'Moldova': 'Eastern Europe',
            'Monaco': 'Western Europe',
            'Montenegro': 'Southern Europe',
            'Netherlands': 'Western Europe',
            'North Macedonia': 'Southern Europe',
            'Norway': 'Northern Europe',
            'Poland': 'Central Europe',
            'Portugal': 'Southern Europe',
            'Romania': 'Eastern Europe',
            'Russia': 'Eastern Europe',  # Transcontinental
            'San Marino': 'Southern Europe',
            'Serbia': 'Southern Europe',
            'Slovakia': 'Central Europe',
            'Slovenia': 'Southern Europe',
            'Spain': 'Southern Europe',
            'Sweden': 'Northern Europe',
            'Switzerland': 'Central Europe',
            'Ukraine': 'Eastern Europe',
            'United Kingdom': 'Northern Europe',
            'Vatican City': 'Southern Europe',

            # Americas
            # North America
            'Canada': 'North America',
            'United States': 'North America',
            'Mexico': 'North America',
            
            # Central America
            'Belize': 'Central America',
            'Costa Rica': 'Central America',
            'El Salvador': 'Central America',
            'Guatemala': 'Central America',
            'Honduras': 'Central America',
            'Nicaragua': 'Central America',
            'Panama': 'Central America',
            
            # Caribbean
            'Antigua and Barbuda': 'Caribbean',
            'Bahamas': 'Caribbean',
            'Barbados': 'Caribbean',
            'Cuba': 'Caribbean',
            'Dominica': 'Caribbean',
            'Dominican Republic': 'Caribbean',
            'Grenada': 'Caribbean',
            'Haiti': 'Caribbean',
            'Jamaica': 'Caribbean',
            'Saint Kitts and Nevis': 'Caribbean',
            'Saint Lucia': 'Caribbean',
            'Saint Vincent and the Grenadines': 'Caribbean',
            'Trinidad and Tobago': 'Caribbean',
            
            # South America
            'Argentina': 'South America',
            'Bolivia': 'South America',
            'Brazil': 'South America',
            'Chile': 'South America',
            'Colombia': 'South America',
            'Ecuador': 'South America',
            'Guyana': 'South America',
            'Paraguay': 'South America',
            'Peru': 'South America',
            'Suriname': 'South America',
            'Uruguay': 'South America',
            'Venezuela': 'South America',

            # Asia
            # Central Asia
            'Kazakhstan': 'Central Asia',
            'Kyrgyzstan': 'Central Asia',
            'Tajikistan': 'Central Asia',
            'Turkmenistan': 'Central Asia',
            'Uzbekistan': 'Central Asia',
            
            # East Asia
            'China': 'East Asia',
            'Japan': 'East Asia',
            'Mongolia': 'East Asia',
            'North Korea': 'East Asia',
            'South Korea': 'East Asia',
            'Taiwan': 'East Asia',
            
            # South Asia
            'Afghanistan': 'South Asia',
            'Bangladesh': 'South Asia',
            'Bhutan': 'South Asia',
            'India': 'South Asia',
            'Maldives': 'South Asia',
            'Nepal': 'South Asia',
            'Pakistan': 'South Asia',
            'Sri Lanka': 'South Asia',
            
            # Southeast Asia
            'Brunei': 'Southeast Asia',
            'Cambodia': 'Southeast Asia',
            'Indonesia': 'Southeast Asia',
            'Laos': 'Southeast Asia',
            'Malaysia': 'Southeast Asia',
            'Myanmar': 'Southeast Asia',
            'Philippines': 'Southeast Asia',
            'Singapore': 'Southeast Asia',
            'Thailand': 'Southeast Asia',
            'Timor-Leste': 'Southeast Asia',
            'Vietnam': 'Southeast Asia',
            
            # Middle East (West Asia)
            'Armenia': 'Middle East',
            'Azerbaijan': 'Middle East',
            'Bahrain': 'Middle East',
            'Cyprus': 'Middle East',
            'Georgia': 'Middle East',
            'Iran': 'Middle East',
            'Iraq': 'Middle East',
            'Israel': 'Middle East',
            'Jordan': 'Middle East',
            'Kuwait': 'Middle East',
            'Lebanon': 'Middle East',
            'Oman': 'Middle East',
            'Palestine': 'Middle East',
            'Qatar': 'Middle East',
            'Saudi Arabia': 'Middle East',
            'Syria': 'Middle East',
            'Turkey': 'Middle East',
            'United Arab Emirates': 'Middle East',
            'Yemen': 'Middle East',

            # ====== AFRICA ====== #
            # Northern Africa
            'Algeria': 'Northern Africa',
            'Egypt': 'Northern Africa',
            'Libya': 'Northern Africa',
            'Morocco': 'Northern Africa',
            'Sudan': 'Northern Africa',
            'Tunisia': 'Northern Africa',
            
            # Sub-Saharan Africa
            # Western Africa
            'Benin': 'Western Africa',
            'Burkina Faso': 'Western Africa',
            'Cape Verde': 'Western Africa',
            "Côte d'Ivoire": 'Western Africa',
            'Gambia': 'Western Africa',
            'Ghana': 'Western Africa',
            'Guinea': 'Western Africa',
            'Guinea-Bissau': 'Western Africa',
            'Liberia': 'Western Africa',
            'Mali': 'Western Africa',
            'Mauritania': 'Western Africa',
            'Niger': 'Western Africa',
            'Nigeria': 'Western Africa',
            'Senegal': 'Western Africa',
            'Sierra Leone': 'Western Africa',
            'Togo': 'Western Africa',
            
            # Central Africa
            'Angola': 'Central Africa',
            'Cameroon': 'Central Africa',
            'Central African Republic': 'Central Africa',
            'Chad': 'Central Africa',
            'Congo (Brazzaville)': 'Central Africa',
            'Congo (Kinshasa)': 'Central Africa',
            'Equatorial Guinea': 'Central Africa',
            'Gabon': 'Central Africa',
            'Sao Tome and Principe': 'Central Africa',
            
            # Eastern Africa
            'Burundi': 'Eastern Africa',
            'Comoros': 'Eastern Africa',
            'Djibouti': 'Eastern Africa',
            'Eritrea': 'Eastern Africa',
            'Ethiopia': 'Eastern Africa',
            'Kenya': 'Eastern Africa',
            'Madagascar': 'Eastern Africa',
            'Malawi': 'Eastern Africa',
            'Mauritius': 'Eastern Africa',
            'Mozambique': 'Eastern Africa',
            'Rwanda': 'Eastern Africa',
            'Seychelles': 'Eastern Africa',
            'Somalia': 'Eastern Africa',
            'South Sudan': 'Eastern Africa',
            'Tanzania': 'Eastern Africa',
            'Uganda': 'Eastern Africa',
            'Zambia': 'Eastern Africa',
            'Zimbabwe': 'Eastern Africa',
            
            # Southern Africa
            'Botswana': 'Southern Africa',
            'Eswatini': 'Southern Africa',
            'Lesotho': 'Southern Africa',
            'Namibia': 'Southern Africa',
            'South Africa': 'Southern Africa',

            # Oceania
            'Australia': 'Australia and New Zealand',
            'New Zealand': 'Australia and New Zealand',
            'Fiji': 'Pacific Islands',
            'Papua New Guinea': 'Pacific Islands',
            'Solomon Islands': 'Pacific Islands',
            'Vanuatu': 'Pacific Islands',

            # Special Regions
            'European Union': 'European Union',
            'African Union': 'African Union',
            'ASEAN': 'ASEAN',
            'Global': 'Global'
        }

    def _load_language_mapping(self) -> Dict[str, str]:
        """Load optimized language to country mapping."""
        # (Same as original)
        return {
            # European Languages
            'dutch': 'Netherlands', 'flemish': 'Belgium',
            'german': 'Germany', 'austrian german': 'Austria',
            'swiss german': 'Switzerland', 'french': 'France',
            'belgian french': 'Belgium', 'swiss french': 'Switzerland',
            'canadian french': 'Canada', 'italian': 'Italy',
            'swiss italian': 'Switzerland', 'spanish': 'Spain',
            'latin american spanish': 'Mexico', 'catalan': 'Spain',
            'portuguese': 'Portugal', 'brazilian portuguese': 'Brazil',
            'english': 'United States', 'british english': 'United Kingdom',
            'irish english': 'Ireland', 'scots': 'United Kingdom',
            'swedish': 'Sweden', 'norwegian': 'Norway',
            'danish': 'Denmark', 'finnish': 'Finland',
            'icelandic': 'Iceland', 'russian': 'Russia',
            'ukrainian': 'Ukraine', 'polish': 'Poland',
            'czech': 'Czech Republic', 'slovak': 'Slovakia',
            'hungarian': 'Hungary', 'romanian': 'Romania',
            'bulgarian': 'Bulgaria', 'serbian': 'Serbia',
            'croatian': 'Croatia', 'slovenian': 'Slovenia',
            'bosnian': 'Bosnia and Herzegovina', 'albanian': 'Albania',
            'greek': 'Greece', 'turkish': 'Turkey',
            'estonian': 'Estonia', 'latvian': 'Latvia',
            'lithuanian': 'Lithuania', 'maltese': 'Malta',

            # Asian Languages
            'mandarin': 'China', 'cantonese': 'China',
            'japanese': 'Japan', 'korean': 'South Korea',
            'vietnamese': 'Vietnam', 'thai': 'Thailand',
            'hindi': 'India', 'bengali': 'Bangladesh',
            'punjabi': 'India', 'tamil': 'India',
            'telugu': 'India', 'urdu': 'Pakistan',
            'indonesian': 'Indonesia', 'malay': 'Malaysia',
            'filipino': 'Philippines', 'burmese': 'Myanmar',
            'khmer': 'Cambodia', 'lao': 'Laos',
            'mongolian': 'Mongolia', 'nepali': 'Nepal',
            'sinhala': 'Sri Lanka', 'dzongkha': 'Bhutan',

            # Middle Eastern Languages
            'arabic': 'Egypt', 'egyptian arabic': 'Egypt',
            'saudi arabic': 'Saudi Arabia', 'hebrew': 'Israel',
            'persian': 'Iran', 'pashto': 'Afghanistan',
            'dari': 'Afghanistan', 'kurdish': 'Iraq',
            'turkmen': 'Turkmenistan', 'uzbek': 'Uzbekistan',
            'kazakh': 'Kazakhstan',

            # African Languages
            'swahili': 'Tanzania', 'amharic': 'Ethiopia',
            'yoruba': 'Nigeria', 'hausa': 'Nigeria',
            'igbo': 'Nigeria', 'somali': 'Somalia',
            'afrikaans': 'South Africa', 'zulu': 'South Africa',
            'xhosa': 'South Africa', 'shona': 'Zimbabwe',
            'malagasy': 'Madagascar',

            # Americas
            'american english': 'United States',
            'canadian english': 'Canada',
            'mexican spanish': 'Mexico',
            'argentine spanish': 'Argentina',
            'brazilian portuguese': 'Brazil',
            'quebec french': 'Canada',
            'haitian creole': 'Haiti',
            'jamaican patois': 'Jamaica',

            # Other
            'esperanto': 'International',
            'latin': 'Vatican City'
        }

    def _load_tld_mapping(self) -> Dict[str, str]:
        """Load optimized TLD to country mapping."""
        # (Same as original)
        return {
            # Europe
            '.al': 'Albania', '.ad': 'Andorra', '.am': 'Armenia', '.at': 'Austria',
            '.az': 'Azerbaijan', '.by': 'Belarus', '.be': 'Belgium', '.ba': 'Bosnia and Herzegovina',
            '.bg': 'Bulgaria', '.hr': 'Croatia', '.cy': 'Cyprus', '.cz': 'Czech Republic',
            '.dk': 'Denmark', '.ee': 'Estonia', '.fi': 'Finland', '.fr': 'France',
            '.ge': 'Georgia', '.de': 'Germany', '.gr': 'Greece', '.hu': 'Hungary',
            '.is': 'Iceland', '.ie': 'Ireland', '.it': 'Italy', '.kz': 'Kazakhstan',
            '.lv': 'Latvia', '.li': 'Liechtenstein', '.lt': 'Lithuania', '.lu': 'Luxembourg',
            '.mk': 'North Macedonia', '.mt': 'Malta', '.md': 'Moldova', '.mc': 'Monaco',
            '.me': 'Montenegro', '.nl': 'Netherlands', '.no': 'Norway', '.pl': 'Poland',
            '.pt': 'Portugal', '.ro': 'Romania', '.ru': 'Russia', '.sm': 'San Marino',
            '.rs': 'Serbia', '.sk': 'Slovakia', '.si': 'Slovenia', '.es': 'Spain',
            '.se': 'Sweden', '.ch': 'Switzerland', '.tr': 'Turkey', '.ua': 'Ukraine',
            '.uk': 'United Kingdom', '.va': 'Vatican City',
           
            # Americas
            '.ai': 'Anguilla', '.ag': 'Antigua and Barbuda', '.ar': 'Argentina', '.aw': 'Aruba',
            '.bs': 'Bahamas', '.bb': 'Barbados', '.bz': 'Belize', '.bm': 'Bermuda',
            '.bo': 'Bolivia', '.br': 'Brazil', '.vg': 'British Virgin Islands', '.ca': 'Canada',
            '.ky': 'Cayman Islands', '.cl': 'Chile', '.co': 'Colombia', '.cr': 'Costa Rica',
            '.cu': 'Cuba', '.dm': 'Dominica', '.do': 'Dominican Republic', '.ec': 'Ecuador',
            '.sv': 'El Salvador', '.fk': 'Falkland Islands', '.gd': 'Grenada', '.gt': 'Guatemala',
            '.gy': 'Guyana', '.ht': 'Haiti', '.hn': 'Honduras', '.jm': 'Jamaica',
            '.mx': 'Mexico', '.ni': 'Nicaragua', '.pa': 'Panama', '.py': 'Paraguay',
            '.pe': 'Peru', '.pr': 'Puerto Rico', '.bl': 'Saint Barthélemy', '.kn': 'Saint Kitts and Nevis',
            '.lc': 'Saint Lucia', '.mf': 'Saint Martin', '.vc': 'Saint Vincent and the Grenadines',
            '.sr': 'Suriname', '.tt': 'Trinidad and Tobago', '.tc': 'Turks and Caicos Islands',
            '.us': 'United States', '.uy': 'Uruguay', '.ve': 'Venezuela',
            
            # Asia
            '.af': 'Afghanistan', '.bh': 'Bahrain', '.bd': 'Bangladesh', '.bt': 'Bhutan',
            '.bn': 'Brunei', '.kh': 'Cambodia', '.cn': 'China', '.tw': 'Taiwan',
            '.in': 'India', '.id': 'Indonesia', '.ir': 'Iran', '.iq': 'Iraq',
            '.il': 'Israel', '.jp': 'Japan', '.jo': 'Jordan', '.kw': 'Kuwait',
            '.kg': 'Kyrgyzstan', '.la': 'Laos', '.lb': 'Lebanon', '.mo': 'Macao',
            '.my': 'Malaysia', '.mv': 'Maldives', '.mn': 'Mongolia', '.mm': 'Myanmar',
            '.np': 'Nepal', '.kp': 'North Korea', '.om': 'Oman', '.pk': 'Pakistan',
            '.ps': 'Palestine', '.ph': 'Philippines', '.qa': 'Qatar', '.sa': 'Saudi Arabia',
            '.sg': 'Singapore', '.kr': 'South Korea', '.lk': 'Sri Lanka', '.sy': 'Syria',
            '.tj': 'Tajikistan', '.th': 'Thailand', '.tl': 'Timor-Leste', '.ae': 'United Arab Emirates',
            '.uz': 'Uzbekistan', '.vn': 'Vietnam', '.ye': 'Yemen',

            # Africa
            '.dz': 'Algeria', '.ao': 'Angola', '.bj': 'Benin', '.bw': 'Botswana',
            '.bf': 'Burkina Faso', '.bi': 'Burundi', '.cv': 'Cape Verde', '.cm': 'Cameroon',
            '.cf': 'Central African Republic', '.td': 'Chad', '.km': 'Comoros', '.cd': 'Congo (Kinshasa)',
            '.cg': 'Congo (Brazzaville)', '.ci': "Côte d'Ivoire", '.dj': 'Djibouti', '.eg': 'Egypt',
            '.gq': 'Equatorial Guinea', '.er': 'Eritrea', '.sz': 'Eswatini', '.et': 'Ethiopia',
            '.ga': 'Gabon', '.gm': 'Gambia', '.gh': 'Ghana', '.gn': 'Guinea',
            '.gw': 'Guinea-Bissau', '.ke': 'Kenya', '.ls': 'Lesotho', '.lr': 'Liberia',
            '.ly': 'Libya', '.mg': 'Madagascar', '.mw': 'Malawi', '.ml': 'Mali',
            '.mr': 'Mauritania', '.mu': 'Mauritius', '.yt': 'Mayotte', '.ma': 'Morocco',
            '.mz': 'Mozambique', '.na': 'Namibia', '.ne': 'Niger', '.ng': 'Nigeria',
            '.re': 'Réunion', '.rw': 'Rwanda', '.sh': 'Saint Helena', '.st': 'Sao Tome and Principe',
            '.sn': 'Senegal', '.sc': 'Seychelles', '.sl': 'Sierra Leone', '.so': 'Somalia',
            '.za': 'South Africa', '.ss': 'South Sudan', '.sd': 'Sudan', '.tz': 'Tanzania',
            '.tg': 'Togo', '.tn': 'Tunisia', '.ug': 'Uganda', '.eh': 'Western Sahara',
            '.zm': 'Zambia', '.zw': 'Zimbabwe',

           # Oceania
            '.as': 'American Samoa', '.au': 'Australia', '.ck': 'Cook Islands', '.fj': 'Fiji',
            '.pf': 'French Polynesia', '.gu': 'Guam', '.ki': 'Kiribati', '.mh': 'Marshall Islands',
            '.fm': 'Micronesia', '.nr': 'Nauru', '.nc': 'New Caledonia', '.nz': 'New Zealand',
            '.nu': 'Niue', '.nf': 'Norfolk Island', '.mp': 'Northern Mariana Islands', '.pw': 'Palau',
            '.pg': 'Papua New Guinea', '.pn': 'Pitcairn', '.ws': 'Samoa', '.sb': 'Solomon Islands',
            '.tk': 'Tokelau', '.to': 'Tonga', '.tv': 'Tuvalu', '.vu': 'Vanuatu',

            # ====== SPECIAL CASES & SECOND-LEVEL DOMAINS ======
            # UK second-level domains
            '.ac.uk': 'United Kingdom', '.gov.uk': 'United Kingdom', '.co.uk': 'United Kingdom',
            '.org.uk': 'United Kingdom', '.me.uk': 'United Kingdom', '.net.uk': 'United Kingdom',
                
            # Australia second-level
            '.com.au': 'Australia', '.org.au': 'Australia', '.net.au': 'Australia',
            '.gov.au': 'Australia', '.edu.au': 'Australia',
                
            # Other notable second-levels
            '.edu': 'United States', '.gov': 'United States', '.mil': 'United States',
            '.gc.ca': 'Canada', '.gob.mx': 'Mexico', '.gov.in': 'India',
                
            # Generic TLDs with geographic significance
            '.asia': 'Asia', '.eu': 'European Union', '.lat': 'Latin America'
        }

    def detect_countries(self, text: str) -> Dict[str, str]:
        """Detect mentioned countries and regions in English text."""
        english_text = self.translate_to_english(text)
        if not english_text:
            return {
                'mentioned_countries': "None",
                'mentioned_regions': "None"
            }
            
        text_lower = english_text.lower()
        mentioned_countries = []
        
        for country, patterns in self.country_regex.items():
            if any(pattern.search(text_lower) for pattern in patterns):
                mentioned_countries.append(country)
        
        # Get unique regions for mentioned countries
        mentioned_regions = list(set(
            self.region_mapping.get(country, "Unknown")
            for country in mentioned_countries
        ))
        
        return {
            'mentioned_countries': ", ".join(mentioned_countries) if mentioned_countries else "None",
            'mentioned_regions': ", ".join(mentioned_regions) if mentioned_regions else "None"
        }

    def infer_country(self, url: str, language: str) -> str:
        """Infer country based on URL and language with improved logic."""
        if not url and not language:
            return "Unknown"
            
        url_lower = url.lower()
        
        # Check TLDs first (prioritizing more specific matches)
        for tld in sorted(self.country_tlds.keys(), key=len, reverse=True):
            if tld in url_lower:
                return self.country_tlds[tld]
        
        # Fallback to language mapping
        if language:
            return self.language_to_country.get(language.lower(), "Unknown")
        
        return "Unknown"

class InfarmedNewsSpider(scrapy.Spider):
    """Optimized spider for scraping news from Infarmed website with OpenPyXL export."""
    
    name = 'infarmed7'
    allowed_domains = ['infarmed.pt']
    start_urls = ['https://www.infarmed.pt/web/infarmed/noticias']
    
    custom_settings = {
        'DOWNLOAD_DELAY': 1.5,
        'USER_AGENT': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'CONCURRENT_REQUESTS': 4,
        'ROBOTSTXT_OBEY': True,
        'RETRY_TIMES': 3,
        'RETRY_HTTP_CODES': [500, 502, 503, 504, 522, 524, 408, 429],
        'HTTPCACHE_ENABLED': True
    }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._initialize_nltk()
        self.classifier = DocumentClassifier()
        self.country_detector = CountryDetector()
        self.summarizer = LsaSummarizer()
        self.summary_sentences = 3
        self.fasttext_model = self._load_fasttext_model()
        self.exporter = ExcelExporter()
        self.page_count = 0 
        
    def closed(self, reason):
        """Called when the spider is closed to save the Excel file."""
        self.exporter.save()
        
    def _initialize_nltk(self):
        """Initialize NLTK resources with error handling."""
        try:
            import nltk
            try:
                nltk.data.find('tokenizers/punkt')
            except LookupError:
                nltk.download('punkt', quiet=True)
        except Exception as e:
            logging.warning(f"NLTK initialization failed: {e}")

    def _load_fasttext_model(self):
        """Load FastText language detection model with improved error handling."""
        try:
            model_path = os.path.join(os.path.dirname(__file__), 'lid.176.ftz')
            if os.path.exists(model_path):
                return fasttext.load_model(model_path)
            
            # Try downloading if not found locally
            import requests
            url = "https://dl.fbaipublicfiles.com/fasttext/supervised-models/lid.176.ftz"
            response = requests.get(url, stream=True)
            if response.status_code == 200:
                with open(model_path, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        f.write(chunk)
                return fasttext.load_model(model_path)
                
            logging.warning("FastText model download failed")
            return None
        except Exception as e:
            logging.warning(f"Fasttext model error: {e}")
            return None

    def detect_language(self, text: str, context_url: str = "") -> str:
        """More robust language detection with fallbacks."""
        if not text or len(text.strip()) < 3:
            return "Unknown"
            
        # Check for Portuguese context first
        if 'infarmed.pt' in context_url.lower():
            return 'Portuguese'
            
        try:
            # Try FastText first
            if self.fasttext_model:
                prediction = self.fasttext_model.predict(text.replace("\n", " "))
                lang_code = prediction[0][0].replace('__label__', '')
                try:
                    lang_name = Lang(lang_code).name
                    # Handle special cases
                    if lang_name.lower() == 'portuguese (brazil)':
                        return 'Portuguese'
                    return lang_name
                except:
                    pass  # Fall through to langdetect
        
            # Fallback to langdetect
            lang_code = detect(text)
            lang_name = Lang(alpha2=lang_code).name
            
            # Special case for Portuguese
            if lang_name.lower() == 'portuguese (brazil)':
                return 'Portuguese'
            return lang_name
        except Exception as e:
            logging.debug(f"Language detection failed: {str(e)}")
            return "Unknown"

    def generate_summary(self, text: str, language: str = 'english') -> str:
        """Generate better quality summaries with context awareness."""
        if not text.strip():
            return ""
        text = re.sub(r'\[.*?\]|\{.*?\}', '', text)  # Remove brackets content
        text = re.sub(r'\s+', ' ', text).strip()
        
        if len(text.split()) <= 15:
            return text
        try:
            try:
                tokenizer = Tokenizer(language)
                parser = PlaintextParser.from_string(text, tokenizer)
                summarizer = LsaSummarizer()
                summary = summarizer(parser.document, self.summary_sentences)
                summary_text = " ".join(str(s) for s in summary)
                
            except:
                tokenizer = Tokenizer('english')
                parser = PlaintextParser.from_string(text, tokenizer)
                summarizer = LsaSummarizer()
                summary = summarizer(parser.document, self.summary_sentences)
                summary_text = " ".join(str(s) for s in summary)
            
            return summary_text[:600].rsplit(' ', 1)[0] + ("..." if len(summary_text) > 600 else "")
        except Exception as e:
            logging.warning(f"Summarization failed: {str(e)}")
            return text[:550].rsplit(' ', 1)[0] + "..."
                
    def parse(self, response: Response):
        """Parse the main news listing page with improved error handling."""
        try:
            self.page_count += 1
            articles = response.css('article.event')
            if not articles:
                self.logger.warning(f"No articles found on page: {response.url}")
                return
            for article in articles:
                try:
                    title = article.css('h4.title::text').get(default="").strip()
                    if not title:
                        continue
                    
                    date = article.xpath('normalize-space(.//p[strong="Data"]/text()[2])').get()
                    if date:
                        date = date.strip().replace(': ', '')
                        
                    detail_url = article.css('a.event-link::attr(href)').get()
                    if not detail_url:
                        self.logger.debug(f"No detail URL found for article: {title}")
                        continue
                    
                    detail_url = response.urljoin(detail_url)
                    lang = self.detect_language(title, detail_url)
                    
                    news_item = {
                        'original_title': title,
                        'english_title': self.classifier.translate_text(title, lang),
                        'date': date,
                        'url': detail_url,
                        'language': lang,
                        'regulatory_body': 'INFARMED',
                        'source_url': response.url
                        }
                    
                    yield response.follow(
                        detail_url,
                        callback=self.parse_article,
                        cb_kwargs={'main_item': news_item},
                        errback=self.handle_error,
                        priority=1
                        )
                    
                except Exception as e:
                    self.logger.error(f"Error processing article: {str(e)}", exc_info=True)
                    
            if self.page_count < 3:            
                next_page_url = response.css('a.link-next::attr(href)').get()
                if next_page_url and next_page_url.lower() != 'javascript:;':
                    try:
                        yield response.follow(
                            next_page_url, 
                            callback=self.parse,
                            errback=self.handle_error
                            )
                    except ValueError as e:
                        self.logger.warning(f"Invalid pagination URL: {next_page_url}. Error: {e}")
        except Exception as e:
            self.logger.error(f"Error parsing page {response.url}: {str(e)}", exc_info=True)

    def parse_article(self, response: Response, main_item: Dict[str, Any]) -> Generator[Dict[str, Any], None, None]:
        """Parse article detail page and enrich the news item."""
        try:
            content_paragraphs = response.css('div.journal-content-article p::text, div.article-body p::text').getall()
            content = ' '.join(p.strip() for p in content_paragraphs if p.strip())
            if not content:
                content = ' '.join(response.xpath('//div[contains(@class, "content")]//text()').getall()).strip()
            content = re.sub(r'\s+', ' ', content).strip()
            english_content = self.classifier.translate_text(content, main_item['language'])
            # Original scraped text
            original_content = content

            # Translate for classification/summarization
            english_content = self.classifier.translate_text(original_content, main_item['language'])

            # Extract drug names BEFORE translation
            drug_names = self.classifier.extract_drug_names(original_content)

            # Use drug_names in classify_product
            product_class = self.classifier.classify_product(english_content)
            product_class['drug_names'] = ", ".join(drug_names) if drug_names else None

            
            english_content = self.classifier.translate_text(content, main_item['language'])
            
            summary = self.generate_summary(content, main_item['language'])
            english_summary = self.classifier.translate_text(summary, main_item['language'])
            
            doc_class = self.classifier.classify_document(english_content)
            product_class = self.classifier.classify_product(english_content)
            countries = self.country_detector.detect_countries(english_content)
            
            item = {
                'title': main_item.get('english_title', ''),
                'summary': english_summary,
                'article_url': main_item.get('url', ''),
                'date': main_item.get('date', ''),
                'document_type': doc_class.get('document_type', ''),
                'product_type': product_class.get('product_type', ''),
                'countries': countries.get('mentioned_countries', ''),
                'regions': countries.get('mentioned_regions', ''),
                'drug_names': product_class.get('drug_names', ''),
                'language': main_item.get('language', ''),
                'source_url': main_item.get('source_url', '') 
            }


            
            # Add item to Excel exporter
            self.exporter.add_item(item)
            
            yield item
         
        except Exception as e:
            self.logger.error(f"Error parsing article {response.url}: {str(e)}", exc_info=True)
            yield main_item  # Return at least the basic info if processing fails
        
    def _extract_content(self, response) -> str:
        """Improved content extraction for INFARMED website."""
        content = response.xpath('''
                                 //div[contains(@class, "journal-content-article")]//p[not(@class)]//text()|
                                 //div[contains(@class, "content")]//p[not(@class)]//text()|
                                 //article//p[not(@class)]//text()
                                 ''').getall()
        if not content:
            content = response.xpath('//body//text()[not(ancestor::script)]').getall()
            filtered = []
            for text in content:
                text = re.sub(r'\s+', ' ', text).strip()
                if (len(text) > 30 and 
                    not text.startswith(('©', 'function(', 'var ', 'window.')) and
                    not re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', text)):
                    filtered.append(text)
                    return ' '.join(filtered) if filtered else ""

    def _fallback_content_extraction(self, response) -> str:
        """Fallback content extraction method."""
        
        selectors = [
            '//div[contains(@class, "body")]//text()',
            '//div[contains(@class, "text")]//text()',
            '//div[contains(@class, "main")]//text()',
            '//div[@role="main"]//text()',
            '//div[contains(@class, "article")]//text()'
            ]
        
        for selector in selectors:
            texts = response.xpath(selector).getall()
            if texts:
                cleaned = [t.strip() for t in texts if t.strip()]
                if cleaned:
                    return ' '.join(cleaned)
                return ""
            
    def handle_error(self, failure):
        """Handle request errors gracefully."""
        self.logger.error(f"Request failed: {failure.value}")


if __name__ == "__main__":
    process = CrawlerProcess()
    process.crawl(InfarmedNewsSpider) 
    process.start()

